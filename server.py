import http.server
import socketserver
import json
import os
import sys
import urllib.parse
from openpyxl import Workbook, load_workbook
import datetime

PORT = 8080
EXCEL_FILE = os.path.join(os.getcwd(), "warranty_registry_backup.xlsx")
PENDING_FILE = os.path.join(os.getcwd(), "pending_backup.json")
PENDING_CARD_FILE = os.path.join(os.getcwd(), "pending_card_updates.json")

RECORDS_CACHE = []
PENDING_WRITES = []
PENDING_CARD_UPDATES = []

def init_excel():
    """Ensure the Excel file exists with proper headers."""
    if not os.path.exists(EXCEL_FILE):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Warranty Registry"
            headers = [
                "Customer Name", 
                "Phone Number", 
                "Service Address", 
                "Product Model", 
                "Brand Name", 
                "Serial Number", 
                "Purchase Date", 
                "Warranty Duration (Months)", 
                "Backup Timestamp",
                "Warranty Card Given"
            ]
            ws.append(headers)
            wb.save(EXCEL_FILE)
            print("Initialized new Excel file.")
        except Exception as e:
            print(f"Error initializing Excel file: {e}")

def load_cache_from_excel_and_json():
    """Load records from Excel and pending JSON cache into memory cache."""
    global RECORDS_CACHE, PENDING_WRITES, PENDING_CARD_UPDATES
    init_excel()
    
    # 1. Read from Excel
    excel_records = []
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                for row in rows[1:]:
                    if not row[0] and not row[5]:
                        continue
                    # Read Excel row columns. Column J (index 9) represents Warranty Card Given
                    excel_records.append({
                        "name": row[0],
                        "phone": str(row[1]) if row[1] is not None else "",
                        "address": row[2],
                        "product": row[3],
                        "brand": row[4],
                        "serial": str(row[5]) if row[5] is not None else "",
                        "date": row[6],
                        "duration": str(row[7]) if row[7] is not None else "24",
                        "cardGiven": str(row[9]) if len(row) > 9 and row[9] is not None else "No"
                    })
            print(f"Loaded {len(excel_records)} records from Excel.")
    except Exception as e:
        print(f"Notice: Excel read skipped or locked on startup ({e}).")

    # 2. Read from pending registrations JSON
    json_records = []
    if os.path.exists(PENDING_FILE):
        try:
            with open(PENDING_FILE, 'r', encoding='utf-8') as f:
                json_records = json.load(f)
                if not isinstance(json_records, list):
                    json_records = []
            print(f"Loaded {len(json_records)} pending registrations from JSON cache.")
        except Exception as e:
            print(f"Error reading JSON backup: {e}")

    # 3. Read from pending card updates JSON
    if os.path.exists(PENDING_CARD_FILE):
        try:
            with open(PENDING_CARD_FILE, 'r', encoding='utf-8') as f:
                PENDING_CARD_UPDATES = json.load(f)
                if not isinstance(PENDING_CARD_UPDATES, list):
                    PENDING_CARD_UPDATES = []
            print(f"Loaded {len(PENDING_CARD_UPDATES)} pending card updates from JSON cache.")
        except Exception as e:
            print(f"Error reading pending cards: {e}")
            PENDING_CARD_UPDATES = []

    # 4. Merge records (allow duplicate serials now!)
    # To determine if a pending record is already saved to Excel, we check all main fields
    excel_signatures = set(
        f"{r['name']}_{r['phone']}_{r['serial']}_{r['date']}".strip().upper() 
        for r in excel_records
    )
    
    PENDING_WRITES = []
    for r in json_records:
        sig = f"{r.get('name', '')}_{r.get('phone', '')}_{r.get('serial', '')}_{r.get('date', '')}".strip().upper()
        if sig not in excel_signatures:
            excel_records.append(r)
            PENDING_WRITES.append(r)
            excel_signatures.add(sig)

    RECORDS_CACHE = excel_records
    
    # 5. Apply pending card updates to RECORDS_CACHE
    for serial in PENDING_CARD_UPDATES:
        serial_upper = serial.strip().upper()
        for r in RECORDS_CACHE:
            if r.get("serial", "").strip().upper() == serial_upper:
                r["cardGiven"] = "Yes"

    print(f"Total active records in server memory cache: {len(RECORDS_CACHE)}")
    
    save_pending_json()
    save_pending_cards_json()
    
    # Flush if unlocked
    if PENDING_WRITES:
        flush_pending_writes()
    if PENDING_CARD_UPDATES:
        flush_card_updates()

def save_pending_json():
    try:
        with open(PENDING_FILE, 'w', encoding='utf-8') as f:
            json.dump(PENDING_WRITES, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving pending JSON: {e}")

def save_pending_cards_json():
    try:
        with open(PENDING_CARD_FILE, 'w', encoding='utf-8') as f:
            json.dump(PENDING_CARD_UPDATES, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving pending card updates JSON: {e}")

def flush_pending_writes():
    """Attempt to write queued new registrations to Excel (allowing infinite duplicate serials)."""
    global PENDING_WRITES
    if not PENDING_WRITES:
        return True
        
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Load existing signatures in Excel to avoid writing identical duplicates on startup/sync
        file_signatures = set()
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for row in rows[1:]:
                # name_phone_serial_date
                sig = f"{row[0]}_{row[1]}_{row[5]}_{row[6]}".strip().upper()
                file_signatures.add(sig)
                    
        successful_records = []
        for record in PENDING_WRITES:
            sig = f"{record.get('name', '')}_{record.get('phone', '')}_{record.get('serial', '')}_{record.get('date', '')}".strip().upper()
            
            # Append if this specific record is not already in Excel
            if sig not in file_signatures:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([
                    record.get("name", ""),
                    record.get("phone", ""),
                    record.get("address", ""),
                    record.get("product", ""),
                    record.get("brand", ""),
                    record.get("serial", ""),
                    record.get("date", ""),
                    int(record.get("duration", 24)) if record.get("duration") else 24,
                    timestamp,
                    record.get("cardGiven", "No")
                ])
                file_signatures.add(sig)
            successful_records.append(record)
            
        wb.save(EXCEL_FILE)
        print(f"Successfully appended {len(successful_records)} entries to Excel.")
        PENDING_WRITES = [r for r in PENDING_WRITES if r not in successful_records]
        save_pending_json()
        return True
    except PermissionError:
        print("Excel sheet is LOCKED (can't append). Retaining in memory/JSON backup.")
        return False
    except Exception as e:
        print(f"Error during Excel append flush: {e}")
        return False

def flush_card_updates():
    """Attempt to apply pending locked card status updates to Excel."""
    global PENDING_CARD_UPDATES
    if not PENDING_CARD_UPDATES:
        return True
        
    init_excel()
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        successful_serials = []
        for serial in PENDING_CARD_UPDATES:
            serial_upper = serial.strip().upper()
            updated = False
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue
                # Match serial number. If duplicates exist, it updates all matching serial rows!
                if row[5] and str(row[5]).strip().upper() == serial_upper:
                    ws.cell(row=row_idx, column=10).value = "Yes"
                    updated = True
            if updated:
                successful_serials.append(serial)
                
        wb.save(EXCEL_FILE)
        if successful_serials:
            print(f"Successfully updated Excel card status for serials: {successful_serials}")
            
        PENDING_CARD_UPDATES = [s for s in PENDING_CARD_UPDATES if s not in successful_serials]
        save_pending_cards_json()
        return True
    except PermissionError:
        print("Excel sheet is LOCKED (can't update card status). Retaining updates in memory/JSON.")
        return False
    except Exception as e:
        print(f"Error flushing card updates to Excel: {e}")
        return False

class BackupHandler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        if self.path.startswith('/api/'):
            self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
            self.send_header('Pragma', 'no-cache')
            self.send_header('Expires', '0')
        super().end_headers()

    def do_GET(self):
        if self.path == '/api/warranties':
            try:
                flush_pending_writes()
                flush_card_updates()
                
                response = json.dumps(RECORDS_CACHE).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Length', len(response))
                self.end_headers()
                self.wfile.write(response)
            except Exception as e:
                self.send_error_response(500, f"Failed to get warranties: {str(e)}")
                
        elif self.path == '/api/open-excel':
            try:
                flush_pending_writes()
                flush_card_updates()
                
                if os.path.exists(EXCEL_FILE):
                    if hasattr(os, 'startfile'):
                        os.startfile(EXCEL_FILE)
                        msg = "Excel backup opened successfully in system editor."
                        status = "success"
                    else:
                        msg = "System does not support direct file launching (Windows-only command)."
                        status = "unsupported"
                else:
                    msg = "Excel backup file does not exist yet."
                    status = "error"
                    
                response = json.dumps({"status": status, "message": msg}).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Length', len(response))
                self.end_headers()
                self.wfile.write(response)
            except Exception as e:
                self.send_error_response(500, f"Failed to open Excel: {str(e)}")
                
        elif self.path == '/api/excel-path':
            try:
                response = json.dumps({"path": EXCEL_FILE}).encode('utf-8')
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Length', len(response))
                self.end_headers()
                self.wfile.write(response)
            except Exception as e:
                self.send_error_response(500, f"Failed to get Excel path: {str(e)}")
        else:
            super().do_GET()

    def do_POST(self):
        if self.path == '/api/warranty':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                record = json.loads(post_data.decode('utf-8'))
                
                if not record.get("serial") or not record.get("name"):
                    self.send_error_response(400, "Missing required parameters: serial or name.")
                    return
                
                serial = str(record.get("serial", "")).strip()
                name = str(record.get("name", "")).strip()
                
                if "cardGiven" not in record:
                    record["cardGiven"] = "No"
                
                # Append directly (allowing duplicate serials!)
                RECORDS_CACHE.append(record)
                PENDING_WRITES.append(record)
                save_pending_json()
                print(f"Added to server cache (boundless): {name} | {serial}")
                
                # Flush
                flushed = flush_pending_writes()
                
                msg = "Warranty successfully recorded."
                if not flushed:
                    msg = "Warranty recorded in server cache. Excel sheet is locked and will sync when unlocked."
                    
                response = json.dumps({
                    "status": "success", 
                    "message": msg
                }).encode('utf-8')
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Length', len(response))
                self.end_headers()
                self.wfile.write(response)
            except Exception as e:
                self.send_error_response(500, f"Failed to write warranty registration: {str(e)}")
                
        elif self.path == '/api/warranty/card-given':
            try:
                content_length = int(self.headers['Content-Length'])
                post_data = self.rfile.read(content_length)
                req_data = json.loads(post_data.decode('utf-8'))
                
                serial = str(req_data.get("serial", "")).strip()
                if not serial:
                    self.send_error_response(400, "Missing required parameter: serial.")
                    return
                
                serial_upper = serial.upper()
                updated_in_cache = False
                
                # Update all matching serials in cache (since duplicates are allowed)
                for r in RECORDS_CACHE:
                    if r.get("serial", "").strip().upper() == serial_upper:
                        r["cardGiven"] = "Yes"
                        updated_in_cache = True
                
                if not updated_in_cache:
                    self.send_error_response(404, f"Serial number {serial} not found in database.")
                    return
                
                # Queue card update
                if serial not in PENDING_CARD_UPDATES:
                    PENDING_CARD_UPDATES.append(serial)
                    save_pending_cards_json()
                
                # Attempt Excel write
                flushed = flush_card_updates()
                
                msg = "Warranty card status successfully updated to Handed Over."
                if not flushed:
                    msg = "Warranty card status saved in server memory cache. Excel file is locked and will sync when unlocked."
                    
                response = json.dumps({
                    "status": "success",
                    "message": msg
                }).encode('utf-8')
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Content-Length', len(response))
                self.end_headers()
                self.wfile.write(response)
            except Exception as e:
                self.send_error_response(500, f"Failed to update card status: {str(e)}")
        else:
            self.send_response(404)
            self.end_headers()

    def send_error_response(self, code, message):
        response = json.dumps({"status": "error", "message": message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', len(response))
        self.end_headers()
        self.wfile.write(response)

if __name__ == '__main__':
    load_cache_from_excel_and_json()
    
    print(f"==================================================")
    print(f"PROBALAJI AI Enterprise Server Booted (V4 Boundless)")
    print(f"Local Server Port: {PORT}")
    print(f"Web portal:        http://localhost:{PORT}")
    print(f"Backup Sheet path: {EXCEL_FILE}")
    print(f"JSON Cache path:   {PENDING_FILE}")
    print(f"Card Cache path:   {PENDING_CARD_FILE}")
    print(f"==================================================")
    
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), BackupHandler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nShutting down server...")
            sys.exit(0)
